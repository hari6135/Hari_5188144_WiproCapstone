import os

from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):

        data = []

        # project root path
        base_dir = os.path.dirname(os.path.dirname(__file__))

        # full excel file path
        file_path = os.path.join(
            base_dir,
            "testdata",
            file_name
        )

        # load workbook
        workbook = load_workbook(file_path)

        # select sheet
        sheet = workbook[sheet_name]

        # read headers from first row
        headers = [cell.value for cell in sheet[1]]

        # read all rows from second row
        for row in sheet.iter_rows(min_row=2, values_only=True):

            # convert row into dictionary
            row_data = dict(zip(headers, row))

            # add to list
            data.append(row_data)

        workbook.close()

        return data