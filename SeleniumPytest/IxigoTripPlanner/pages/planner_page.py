import time

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from utils.logger import LogGen


logger = LogGen.loggen()


class PlannerPage:

    def __init__(self, driver):
        self.driver = driver

    def click_plan(self):

        logger.info("Waiting for Plan button")

        plan_btn = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,"//a[@data-testid='submenu-item' and contains(@href,'plan')]")))

        assert plan_btn.is_displayed(), "Plan button is not visible"

        logger.info("Clicking Plan button")

        plan_btn.click()

        logger.info("Plan page opened")

        time.sleep(2)


    #travel month selection
    def select_travel_month(self):
        logger.info("Waiting for Travel Month section")

        month_btn = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//span[contains(text(),'Travel month')]"
                )
            )
        )

        month_btn.click()

        logger.info("Travel Month clicked")

        time.sleep(2)

        logger.info("Waiting for August option")

        august = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//span[text()='August']"
                )
            )
        )

        august.click()

        logger.info("August selected")

        time.sleep(2)


    # click from place
    def click_from_location(self):

        logger.info("Waiting for From place field")

        from_place = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//span[@class='body-xs' and contains(text(),'From')]"
                )
            )
        )

        assert from_place.is_displayed(), "From place field is not visible"

        from_place.click()

        logger.info("From place field clicked")

    def enter_from_location(self):
        logger.info("Opening Excel file")

        workbook = load_workbook("testdata/places.xlsx")

        sheet = workbook["Sheet1"]

        from_location = sheet.cell(row=2, column=1).value

        logger.info(f"From Location: {from_location}")

        logger.info("Waiting for search input")

        from_search = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//input[@placeholder='Search']"
                )
            )
        )

        from_search.click()

        from_search.clear()

        from_search.send_keys(from_location)

        time.sleep(2)

        logger.info("From location entered")


    # click the first locatio from list
    def click_from_list(self):

        logger.info("Waiting for From place field")

        from_place = WebDriverWait(self.driver, 20).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,"(//div[@class='overflow-auto']//a)[1]") ) )

        from_place.click()

        logger.info("Clicked the location from the list")

        time.sleep(2)